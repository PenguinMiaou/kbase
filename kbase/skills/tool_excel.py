"""Excel tools — read structure, write cells with formatting, create workbooks."""
import json
import shutil
from pathlib import Path
from .tools import SkillTool, register_tool


class ExcelReadTool(SkillTool):
    name = "excel_read"
    description = "Read Excel file structure: headers, row count, specific rows/cells, or a range."
    is_read_only = True
    max_result_chars = 12000
    input_schema = {
        "type": "object",
        "properties": {
            "file_path": {"type": "string", "description": "Path to .xlsx/.xls file"},
            "sheet": {"type": "string", "description": "Sheet name (default: active sheet)"},
            "range": {"type": "string", "description": "Cell range like A1:D10, or 'headers' for first row, or 'all' for full dump (max 200 rows)"},
            "row_start": {"type": "integer", "description": "Start row (1-based)"},
            "row_end": {"type": "integer", "description": "End row (1-based, max 200 rows from start)"},
        },
        "required": ["file_path"],
    }

    def call(self, params: dict) -> str:
        import openpyxl
        fp = params["file_path"]
        if not Path(fp).exists():
            return f"Error: File not found: {fp}"

        wb = openpyxl.load_workbook(fp, data_only=True)
        sheet_name = params.get("sheet") or wb.sheetnames[0]
        if sheet_name not in wb.sheetnames:
            return f"Error: Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
        ws = wb[sheet_name]

        result = {
            "file": Path(fp).name,
            "sheets": wb.sheetnames,
            "active_sheet": sheet_name,
            "dimensions": ws.dimensions,
            "max_row": ws.max_row,
            "max_col": ws.max_column,
        }

        range_param = params.get("range", "headers")

        if range_param == "headers":
            headers = []
            for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=False), []):
                try:
                    headers.append({"col": cell.column_letter, "value": cell.value})
                except AttributeError:
                    pass  # skip merged cells
            result["headers"] = headers

        elif range_param == "all" or (params.get("row_start") and params.get("row_end")):
            r_start = params.get("row_start", 1)
            r_end = min(params.get("row_end", ws.max_row), r_start + 199)
            rows = []
            for row in ws.iter_rows(min_row=r_start, max_row=r_end, values_only=False):
                row_data = {}
                for cell in row:
                    try:
                        if cell.value is not None:
                            row_data[cell.column_letter] = str(cell.value)
                    except AttributeError:
                        pass  # skip merged cells
                if row_data:
                    rows.append({"row": row[0].row, "data": row_data})
            result["rows"] = rows
            result["row_range"] = f"{r_start}-{r_end}"

        else:
            # Specific range like A1:D10
            rows = []
            for row in ws[range_param]:
                row_data = {}
                for cell in row:
                    if cell.value is not None:
                        row_data[cell.coordinate] = str(cell.value)
                if row_data:
                    rows.append(row_data)
            result["cells"] = rows

        wb.close()
        return json.dumps(result, ensure_ascii=False, indent=2)


class ExcelWriteTool(SkillTool):
    name = "excel_write"
    description = "Write values to cells in an Excel file. Supports red/blue font for change tracking and cell comments."
    is_read_only = False
    input_schema = {
        "type": "object",
        "properties": {
            "file_path": {"type": "string", "description": "Path to the OUTPUT .xlsx file (will be created/modified)"},
            "sheet": {"type": "string", "description": "Sheet name (default: active)"},
            "changes": {
                "type": "array",
                "description": "List of cell changes",
                "items": {
                    "type": "object",
                    "properties": {
                        "cell": {"type": "string", "description": "Cell reference like A1, B5"},
                        "value": {"type": "string", "description": "New value"},
                        "color": {"type": "string", "enum": ["red", "blue", "black"], "description": "Font color: red=modified, blue=new, black=unchanged"},
                        "comment": {"type": "string", "description": "Source/reason for the change"},
                    },
                    "required": ["cell", "value"],
                },
            },
        },
        "required": ["file_path", "changes"],
    }

    def call(self, params: dict) -> str:
        import openpyxl
        from openpyxl.styles import Font
        from openpyxl.comments import Comment

        fp = params["file_path"]
        if Path(fp).exists():
            wb = openpyxl.load_workbook(fp)
        else:
            wb = openpyxl.Workbook()

        sheet_name = params.get("sheet") or wb.sheetnames[0]
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
        ws = wb[sheet_name]

        color_map = {"red": "FF0000", "blue": "0000FF", "black": "000000"}
        changed = 0

        for change in params.get("changes", []):
            cell_ref = change["cell"]
            cell = ws[cell_ref]
            old_value = cell.value
            cell.value = change["value"]

            color = change.get("color", "red" if old_value else "blue")
            cell.font = Font(color=color_map.get(color, "FF0000"))

            comment_text = change.get("comment", "")
            if old_value and old_value != change["value"]:
                comment_text = f"Original: {old_value}\n{comment_text}".strip()
            if comment_text:
                cell.comment = Comment(comment_text, "KBase")

            changed += 1

        wb.save(fp)
        wb.close()
        return f"OK: {changed} cells written to {Path(fp).name}"


class ExcelCopyTool(SkillTool):
    name = "excel_copy"
    description = "Copy an Excel file to create a working copy. Always do this before modifying — never modify the original."
    is_read_only = False
    input_schema = {
        "type": "object",
        "properties": {
            "source": {"type": "string", "description": "Original file path"},
            "destination": {"type": "string", "description": "New file path for the working copy"},
        },
        "required": ["source", "destination"],
    }

    def call(self, params: dict) -> str:
        src = Path(params["source"])
        dst = Path(params["destination"])
        if not src.exists():
            return f"Error: Source not found: {src}"
        dst.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(str(src), str(dst))
        return f"OK: Copied {src.name} → {dst.name}"


# Register
register_tool(ExcelReadTool())
register_tool(ExcelWriteTool())
register_tool(ExcelCopyTool())
