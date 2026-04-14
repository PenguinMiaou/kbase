"""Skill Harness — query loop engine that orchestrates LLM + tools.

Two execution modes:
  1. NATIVE tool_use (Claude/GPT-4): LLM calls tools directly via API
  2. GUIDED pipeline (any LLM): Deterministic pipeline, LLM generates content at each step

The harness handles: validation → execution → result budgeting → error recovery.
"""
import json
import time
import asyncio
import queue
import threading
from pathlib import Path
from typing import Callable, Optional

from .tools import get_tool, get_all_tools, get_tool_schemas

# Import tool modules to ensure registration
from . import tool_excel, tool_search, tool_file


class SkillHarness:
    """Execute a skill by orchestrating LLM and tools."""

    def __init__(self, llm_func: Callable, progress_queue: queue.Queue = None):
        self.llm = llm_func
        self.progress = progress_queue
        self._stop = False

    def stop(self):
        self._stop = True

    def emit(self, event_type: str, data: dict):
        if self.progress:
            self.progress.put(json.dumps({"type": event_type, **data}))

    # ── Guided Pipeline Mode (works with any LLM) ────────────────

    def run_document_enrich(self, instruction: str, file_path: str,
                            use_kb: bool = True, use_web: bool = True) -> dict:
        """Guided pipeline for document enrichment. Works with any LLM."""
        start = time.time()
        fp = Path(file_path)
        ext = fp.suffix.lower()

        self.emit("status", {"message": f"Reading {fp.name}...", "step": "parse"})

        # Step 1: Parse file structure
        if ext in (".xlsx", ".xls"):
            return self._enrich_excel(instruction, file_path, use_kb, use_web)
        elif ext == ".pptx":
            return self._enrich_pptx(instruction, file_path, use_kb, use_web)
        elif ext == ".docx":
            return self._enrich_docx(instruction, file_path, use_kb, use_web)
        else:
            return {"error": f"Unsupported file type: {ext}. Supported: xlsx, xls, pptx, docx"}

    def _enrich_excel(self, instruction: str, file_path: str,
                      use_kb: bool, use_web: bool) -> dict:
        """Excel enrichment pipeline."""
        fp = Path(file_path)

        # Step 1: Copy original → output file
        output_dir = Path.home() / ".kbase" / "default" / "outputs"
        output_dir.mkdir(parents=True, exist_ok=True)
        ts = time.strftime("%Y%m%d_%H%M%S")
        output_path = output_dir / f"{fp.stem}_enriched_{ts}.xlsx"

        copy_tool = get_tool("excel_copy")
        copy_tool.call({"source": str(fp), "destination": str(output_path)})
        self.emit("status", {"message": "File copied, analyzing structure...", "step": "parse"})

        # Step 2: Read structure
        read_tool = get_tool("excel_read")
        headers_json = read_tool.call({"file_path": str(output_path), "range": "headers"})
        sample_json = read_tool.call({"file_path": str(output_path), "row_start": 1, "row_end": 5, "range": "all"})

        # Step 3: Ask LLM to plan modifications
        self.emit("status", {"message": "Planning modifications...", "step": "plan"})
        plan_prompt = f"""You are a document enrichment expert. The user wants to modify an Excel file.

User instruction: {instruction}

File: {fp.name}
Headers: {headers_json}
Sample data (first 5 rows): {sample_json}

Analyze the file and the instruction. Output a JSON plan:
{{
  "summary": "Brief description of what needs to be done",
  "search_queries": ["query1", "query2", ...],  // Unique search queries needed
  "target_columns": ["A", "B", ...],  // Columns to modify/fill
  "row_range": [start_row, end_row],  // Rows to process (1-based, include header)
  "strategy": "enrich_existing" or "fill_empty" or "add_column" or "restructure"
}}

Output ONLY valid JSON, nothing else."""

        plan_text = self.llm(plan_prompt)
        try:
            # Extract JSON from response
            plan_text = plan_text.strip()
            if "```" in plan_text:
                plan_text = plan_text.split("```")[1].strip()
                if plan_text.startswith("json"):
                    plan_text = plan_text[4:].strip()
            plan = json.loads(plan_text)
        except (json.JSONDecodeError, IndexError):
            plan = {"summary": "General enrichment", "search_queries": [instruction],
                    "target_columns": [], "row_range": [2, 50], "strategy": "enrich_existing"}

        self.emit("plan", {"plan": plan})

        # Step 4: Execute searches
        self.emit("status", {"message": f"Searching {len(plan.get('search_queries', []))} queries...", "step": "search"})
        search_results = {}
        queries = plan.get("search_queries", [instruction])

        for i, query in enumerate(queries[:20]):
            if self._stop:
                break
            self.emit("search", {"query": query, "current": i + 1, "total": len(queries)})

            results_text = ""
            if use_kb:
                kb_tool = get_tool("kb_search")
                kb_result = kb_tool.call({"query": query, "top_k": 3})
                results_text += f"Knowledge Base:\n{kb_result}\n\n"

            if use_web:
                web_tool = get_tool("web_search")
                web_result = web_tool.call({"query": query, "max_results": 3})
                results_text += f"Web Search:\n{web_result}\n\n"

            search_results[query] = results_text

        # Step 5: Read full data and enrich row by row
        import openpyxl
        wb = openpyxl.load_workbook(str(output_path))
        ws = wb.active
        max_row = ws.max_row
        row_start = plan.get("row_range", [2, max_row])[0]
        row_end = min(plan.get("row_range", [2, max_row])[-1], max_row)

        self.emit("status", {"message": f"Enriching rows {row_start}-{row_end}...", "step": "enrich"})

        # Get headers for context
        headers = {}
        for cell in next(ws.iter_rows(min_row=1, max_row=1), []):
            headers[cell.column_letter] = cell.value

        changes_count = 0
        all_search_context = "\n\n".join(f"[Query: {q}]\n{r}" for q, r in search_results.items())

        # Process in batches of 10 rows
        batch_size = 10
        for batch_start in range(row_start, row_end + 1, batch_size):
            if self._stop:
                break
            batch_end = min(batch_start + batch_size - 1, row_end)

            # Read batch
            batch_data = []
            for row in ws.iter_rows(min_row=batch_start, max_row=batch_end, values_only=False):
                row_dict = {"_row": row[0].row}
                for cell in row:
                    if cell.value is not None:
                        row_dict[cell.column_letter] = str(cell.value)
                    else:
                        row_dict[cell.column_letter] = ""
                batch_data.append(row_dict)

            self.emit("enrich", {
                "current_row": batch_start,
                "total_rows": row_end,
                "message": f"Processing rows {batch_start}-{batch_end}..."
            })

            # Ask LLM to enrich this batch
            enrich_prompt = f"""You are enriching an Excel file based on the user's instruction and search results.

User instruction: {instruction}

Column headers: {json.dumps(headers, ensure_ascii=False)}

Current batch (rows {batch_start}-{batch_end}):
{json.dumps(batch_data, ensure_ascii=False, indent=2)}

Search results for reference:
{all_search_context[:6000]}

For each cell that needs modification, output a JSON array:
[
  {{"row": 5, "col": "C", "value": "new value", "type": "modified", "source": "where this info came from"}},
  ...
]

Rules:
- "type" is "modified" (changing existing) or "new" (filling empty cell)
- Always include "source" — the file name or URL where the info came from
- Only include cells that actually need changes
- If no changes needed for this batch, output []

Output ONLY valid JSON array, nothing else."""

            enrich_text = self.llm(enrich_prompt)
            try:
                enrich_text = enrich_text.strip()
                if "```" in enrich_text:
                    enrich_text = enrich_text.split("```")[1].strip()
                    if enrich_text.startswith("json"):
                        enrich_text = enrich_text[4:].strip()
                cell_changes = json.loads(enrich_text)
            except (json.JSONDecodeError, IndexError):
                cell_changes = []

            # Apply changes
            from openpyxl.styles import Font
            from openpyxl.comments import Comment

            for change in cell_changes:
                row_num = change.get("row")
                col = change.get("col")
                new_val = change.get("value")
                change_type = change.get("type", "modified")
                source = change.get("source", "")

                if not row_num or not col or not new_val:
                    continue

                cell = ws[f"{col}{row_num}"]
                old_val = cell.value
                cell.value = new_val

                if change_type == "modified":
                    cell.font = Font(color="FF0000")
                else:
                    cell.font = Font(color="0000FF")

                comment_text = f"Source: {source}"
                if old_val and str(old_val) != str(new_val):
                    comment_text = f"Original: {old_val}\n{comment_text}"
                cell.comment = Comment(comment_text, "KBase")
                changes_count += 1

        # Save
        wb.save(str(output_path))
        wb.close()

        self.emit("status", {"message": "Generating summary...", "step": "summary"})

        # Step 6: Generate summary
        summary = f"Modified {changes_count} cells in {fp.name}. Output: {output_path.name}"

        result = {
            "status": "ok",
            "output_path": str(output_path),
            "output_name": output_path.name,
            "download_url": f"/api/skill/download/{output_path.name}",
            "changes_count": changes_count,
            "summary": summary,
            "elapsed_seconds": round(time.time() - start, 1),
        }
        self.emit("complete", result)
        return result

    def _enrich_pptx(self, instruction: str, file_path: str,
                     use_kb: bool, use_web: bool) -> dict:
        """PPTX enrichment pipeline."""
        import shutil
        fp = Path(file_path)
        output_dir = Path.home() / ".kbase" / "default" / "outputs"
        output_dir.mkdir(parents=True, exist_ok=True)
        ts = time.strftime("%Y%m%d_%H%M%S")
        output_path = output_dir / f"{fp.stem}_enriched_{ts}.pptx"
        shutil.copy2(str(fp), str(output_path))

        # Read PPTX structure
        read_tool = get_tool("pptx_read")
        structure = read_tool.call({"file_path": str(output_path), "slides": "all"})
        self.emit("status", {"message": "Analyzing PPTX...", "step": "parse"})

        # Search for enrichment content
        self.emit("status", {"message": "Searching knowledge base...", "step": "search"})
        kb_tool = get_tool("kb_search")
        kb_result = kb_tool.call({"query": instruction, "top_k": 10})
        web_result = ""
        if use_web:
            web_tool = get_tool("web_search")
            web_result = web_tool.call({"query": instruction, "max_results": 5})

        # Ask LLM for slide-level modifications
        self.emit("status", {"message": "Generating modifications...", "step": "enrich"})
        prompt = f"""You are modifying a PPTX presentation.

User instruction: {instruction}

Current PPTX structure:
{structure}

Knowledge base results:
{kb_result}

Web search results:
{web_result}

Output a JSON array of changes to apply. Each change modifies ONE slide:
[
  {{"slide": 1, "notes": "Speaker notes to add for this slide", "source": "data source"}},
  {{"slide": 3, "notes": "Additional context for slide 3", "source": "data source"}},
]

Add speaker notes with enrichment info. Do NOT modify existing slide text unless explicitly asked.
Output ONLY valid JSON array."""

        changes_text = self.llm(prompt)
        try:
            if "```" in changes_text:
                changes_text = changes_text.split("```")[1].strip()
                if changes_text.startswith("json"):
                    changes_text = changes_text[4:].strip()
            changes = json.loads(changes_text)
        except (json.JSONDecodeError, IndexError):
            changes = []

        write_tool = get_tool("pptx_write")
        write_tool.call({"file_path": str(output_path), "changes": changes})

        result = {
            "status": "ok",
            "output_path": str(output_path),
            "output_name": output_path.name,
            "download_url": f"/api/skill/download/{output_path.name}",
            "changes_count": len(changes),
            "elapsed_seconds": round(time.time() - time.time(), 1),
        }
        self.emit("complete", result)
        return result

    def _enrich_docx(self, instruction: str, file_path: str,
                     use_kb: bool, use_web: bool) -> dict:
        """DOCX enrichment pipeline."""
        import shutil
        fp = Path(file_path)
        output_dir = Path.home() / ".kbase" / "default" / "outputs"
        output_dir.mkdir(parents=True, exist_ok=True)
        ts = time.strftime("%Y%m%d_%H%M%S")
        output_path = output_dir / f"{fp.stem}_enriched_{ts}.docx"
        shutil.copy2(str(fp), str(output_path))

        # Read DOCX
        read_tool = get_tool("docx_read")
        structure = read_tool.call({"file_path": str(output_path), "section": "all"})
        self.emit("status", {"message": "Analyzing DOCX...", "step": "parse"})

        # Search
        self.emit("status", {"message": "Searching knowledge base...", "step": "search"})
        kb_tool = get_tool("kb_search")
        kb_result = kb_tool.call({"query": instruction, "top_k": 10})
        web_result = ""
        if use_web:
            web_tool = get_tool("web_search")
            web_result = web_tool.call({"query": instruction, "max_results": 5})

        # Ask LLM for modifications
        self.emit("status", {"message": "Generating modifications...", "step": "enrich"})
        prompt = f"""You are modifying a DOCX document.

User instruction: {instruction}

Current document structure:
{structure}

Knowledge base:
{kb_result}

Web search:
{web_result}

Output a JSON array of changes:
[
  {{"type": "append", "text": "New paragraph text to add at the end"}},
  {{"type": "replace_paragraph", "paragraph_index": 5, "text": "Modified text"}},
]

Rules:
- "replace_paragraph" appends modification in red to existing paragraph
- "append" adds new content at the end in blue
- Include source attribution in the text
Output ONLY valid JSON array."""

        changes_text = self.llm(prompt)
        try:
            if "```" in changes_text:
                changes_text = changes_text.split("```")[1].strip()
                if changes_text.startswith("json"):
                    changes_text = changes_text[4:].strip()
            changes = json.loads(changes_text)
        except (json.JSONDecodeError, IndexError):
            changes = []

        write_tool = get_tool("docx_write")
        write_tool.call({"file_path": str(output_path), "changes": changes})

        result = {
            "status": "ok",
            "output_path": str(output_path),
            "output_name": output_path.name,
            "download_url": f"/api/skill/download/{output_path.name}",
            "changes_count": len(changes),
        }
        self.emit("complete", result)
        return result
